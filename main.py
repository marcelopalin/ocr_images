import pytesseract
from PIL import Image, ImageEnhance, ImageFilter
from pathlib import Path
from openpyxl import Workbook, load_workbook
from loguru import logger

# Configurações dos diretórios de entrada e saída
input_dir = "./Entrada"
output_dir = "./Saida"
processed_dir = "./Saida/Processed"

logger.add("ocr_extractor.log", rotation="1 MB")

def preprocess_image(image_path: str, save_path: Path) -> Image:
    try:
        image = Image.open(image_path)
        # Convert to grayscale
        image = image.convert('L')
        # Apply sharpening filter
        image = image.filter(ImageFilter.SHARPEN)
        # Increase contrast
        enhancer = ImageEnhance.Contrast(image)
        image = enhancer.enhance(2)
        # Increase brightness
        enhancer = ImageEnhance.Brightness(image)
        image = enhancer.enhance(1.5)
        
        # Save the processed image
        image.save(save_path)
        
        logger.info(f"Imagem pré-processada com sucesso {image_path} para {save_path}")
        return image
    except Exception as e:
        logger.error(f"Erro ao pré-processar a imagem {image_path}: {e}")
        return None

def extract_text_from_image(image_path: str, save_path: Path) -> str:
    try:
        image = preprocess_image(image_path, save_path)
        if image is None:
            return ""
        text = pytesseract.image_to_string(image)
        logger.info(f"Texto extraído com sucesso da imagem {image_path}")
        print(f"Texto extraído da imagem {image_path}:\n{text}\n")
        return text
    except Exception as e:
        logger.error(f"Erro ao extrair texto da imagem {image_path}: {e}")
        return ""

def save_to_excel(data: list, output_path: str):
    if Path(output_path).exists():
        wb = load_workbook(output_path)
        ws = wb.active
        print("Arquivo Excel existente, substituindo...")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados"
        ws.append(["Rank", "Nome", "Pontos"])  # Write header if new file
        print("Novo arquivo Excel criado.")

    for player in data:
        ws.append([player['rank'], player['name'], player['points']])

    wb.save(output_path)
    logger.info(f"Resultados salvos com sucesso em {output_path}")

def clean_text(text: str) -> str:
    # Basic text cleaning
    text = text.replace('\n', ' ')
    text = ' '.join(text.split())
    return text

def parse_text(text: str) -> list:
    lines = text.split(' ')
    players = []
    rank = 1

    i = 0
    while i < len(lines):
        # Procurar pontos na linha atual
        if lines[i].replace(',', '').isdigit():
            points = lines[i]
            # Procurar o nome do jogador
            name_parts = []
            j = i - 1
            while j >= 0 and not lines[j].replace(',', '').isdigit():
                name_parts.insert(0, lines[j])
                j -= 1
            if name_parts:
                name = ' '.join(name_parts).strip()
                players.append({
                    'rank': rank,
                    'name': name,
                    'points': points
                })
                rank += 1
        i += 1

    return players

def main():
    input_path = Path(input_dir)
    output_path = Path(output_dir)
    processed_path = Path(processed_dir)
    
    # Ensure output and processed directories exist
    output_path.mkdir(parents=True, exist_ok=True)
    processed_path.mkdir(parents=True, exist_ok=True)

    # List all JPG files in the input directory
    images = list(input_path.glob("*.jpg"))
    logger.info(f"Encontradas {len(images)} imagens")
    print(f"Encontradas {len(images)} imagens")

    all_players = []

    for idx, image in enumerate(images, start=1):
        logger.info(f"Processando imagem: {image}")
        print(f"Processando imagem: {image}")
        
        processed_image_path = processed_path / f"processed_{image.name}"
        text = extract_text_from_image(image, processed_image_path)

        # Save the OCR text result to a file
        text_output = output_path / f"fig{idx}.txt"
        with text_output.open('w') as f:
            f.write(text)

        cleaned_text = clean_text(text)
        players = parse_text(cleaned_text)
        all_players.extend(players)
        for player in players:
            print(f"Jogador processado: {player['name']}")

    # Save all results to Excel
    excel_output = output_path / "resultados.xlsx"
    save_to_excel(all_players, excel_output)

    logger.info("Processamento concluído.")
    print("Processamento concluído.")

if __name__ == "__main__":
    main()
